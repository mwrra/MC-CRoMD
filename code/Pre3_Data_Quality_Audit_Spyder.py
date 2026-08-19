# -*- coding: utf-8 -*-
"""
Pre3_Data_Quality_Audit_Spyder.py

فحص جودة خصائص MC-CRoMD بعد ربط SourceGroupID وإنشاء التقسيم الثابت.

يفحص لكل سيناريو:
1) NaN والخلايا الفارغة
2) +inf و -inf
3) القيم النصية داخل أعمدة الخصائص
4) القيم السالبة غير المنطقية
5) الخصائص الثابتة اعتمادًا على Train فقط
6) القيم المتطرفة بحدود IQR المحسوبة من Train فقط
7) إنشاء تقرير Excel لكل خاصية ولكل سيناريو

السيناريوهات:
- Binary: colon_aca, colon_n
- Three_Class: lung_aca, lung_scc, lung_n
- Five_Class: جميع الفئات الخمس

ملاحظة:
هذا البرنامج للتدقيق فقط، ولا يعدّل القيم ولا ينفذ التعويض أو Scaling.
"""

from __future__ import annotations

import argparse
import json
import math
import platform
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =========================
# الإعدادات الافتراضية
# =========================

DEFAULT_INPUT_FILE = Path(
    r"G:\My Research About Lung Canser\INASS\1-SourceGroupID"
    r"\GroupedSplitResults\32_5Clas\Features_With_Groups_And_Split.xlsx"
)

DEFAULT_OUTPUT_FILE = Path(
    r"G:\My Research About Lung Canser\INASS\2-DataQualityAudit"
    r"\32_Data_Quality_Audit.xlsx"
)

DEFAULT_IQR_FACTOR = 1.5

SCENARIOS: dict[str, tuple[str, ...]] = {
    "Binary": ("colon_aca", "colon_n"),
    "Three_Class": ("lung_aca", "lung_scc", "lung_n"),
    "Five_Class": (
        "colon_aca",
        "colon_n",
        "lung_aca",
        "lung_scc",
        "lung_n",
    ),
}

# الأعمدة التي ليست خصائص رقمية.
METADATA_COLUMNS = {
    "ImgName",
    "FileName",
    "ImagePath",
    "ClassName",
    "label",
    "SourceGroupID",
    "PatchSize",
    "ImageWidth",
    "ImageHeight",
    "ImageSHA256",
    "Split",
    "split",
    "stem",
    "filename",
    "tissue",
    "group_id",
    "local_cluster_label",
    
    # أعمدة أضيفت من LC25000-clean
    "LC25000_ClassName",
    "LC25000_FileName",
    "LC25000_Tissue",
    "Previous_SourceGroupID",
    "_match_stem_LC25000",
    "LC25000_GroupID",
    "LC25000_LocalCluster",
}

VALID_SPLITS = ("Train", "Validation", "Test")


def normalize_split(value: Any) -> str:
    """توحيد كتابة أسماء أقسام البيانات."""
    text = str(value).strip().lower()
    mapping = {
        "train": "Train",
        "training": "Train",
        "validation": "Validation",
        "valid": "Validation",
        "val": "Validation",
        "test": "Test",
        "testing": "Test",
    }
    return mapping.get(text, str(value).strip())


def is_cronbach_feature(column_name: str) -> bool:
    """Cronbach’s Alpha قد يكون سالبًا إحصائيًا، لذلك لا يعد سالبًا غير منطقي."""
    return column_name.lower().startswith("cronbach")


def expected_nonnegative(column_name: str) -> bool:
    """
    خصائص MC-CRoMD الأخرى يفترض أن تكون غير سالبة:
    range, variance, std, mad, quartiles, IQR, CV.
    """
    return not is_cronbach_feature(column_name)


def safe_number(value: Any) -> Any:
    """تحويل القيم الخاصة إلى قيم مناسبة للكتابة في Excel."""
    if value is None:
        return None
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        if np.isnan(value):
            return None
        if np.isposinf(value):
            return "+inf"
        if np.isneginf(value):
            return "-inf"
        return float(value)
    return value


def read_input(path: Path, sheet_name: str | int = 0) -> pd.DataFrame:
    """قراءة Excel أو CSV."""
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(path, sheet_name=sheet_name)
    if suffix == ".csv":
        return pd.read_csv(path)
    raise ValueError(f"Unsupported input format: {path.suffix}")


def detect_feature_columns(frame: pd.DataFrame) -> list[str]:
    """اكتشاف أعمدة الخصائص مع استبعاد الأعمدة التعريفية."""
    feature_columns = [
        str(column)
        for column in frame.columns
        if str(column) not in METADATA_COLUMNS
    ]

    # نستبعد أي أعمدة إضافية معروفة بأنها تعريفية أو ناتجة عن الدمج.
    excluded_prefixes = (
        "Unnamed:",
    )
    feature_columns = [
        column
        for column in feature_columns
        if not column.startswith(excluded_prefixes)
    ]

    if not feature_columns:
        raise ValueError("No feature columns were detected.")

    return feature_columns


def build_overview(
    frame: pd.DataFrame,
    feature_columns: list[str],
    input_file: Path,
    iqr_factor: float,
) -> pd.DataFrame:
    """ملخص عام للملف قبل فصل السيناريوهات."""
    rows: list[dict[str, Any]] = [
        {"Item": "Input file", "Value": str(input_file)},
        {"Item": "Generated at", "Value": datetime.now().isoformat(timespec="seconds")},
        {"Item": "Total rows", "Value": len(frame)},
        {"Item": "Total columns", "Value": len(frame.columns)},
        {"Item": "Feature columns", "Value": len(feature_columns)},
        {"Item": "IQR factor", "Value": iqr_factor},
        {"Item": "Python version", "Value": platform.python_version()},
        {"Item": "pandas version", "Value": pd.__version__},
        {"Item": "NumPy version", "Value": np.__version__},
    ]

    if "ClassName" in frame.columns:
        for class_name, count in frame["ClassName"].value_counts(dropna=False).items():
            rows.append(
                {
                    "Item": f"Class count: {class_name}",
                    "Value": int(count),
                }
            )

    if "Split" in frame.columns:
        for split_name, count in frame["Split"].value_counts(dropna=False).items():
            rows.append(
                {
                    "Item": f"Split count: {split_name}",
                    "Value": int(count),
                }
            )

    return pd.DataFrame(rows)


def make_scenario_summary(
    scenario_name: str,
    scenario_frame: pd.DataFrame,
    feature_count: int,
) -> pd.DataFrame:
    """ملخص أعداد الصفوف والفئات والأقسام داخل السيناريو."""
    rows: list[dict[str, Any]] = [
        {"Metric": "Scenario", "Value": scenario_name},
        {"Metric": "Rows", "Value": len(scenario_frame)},
        {"Metric": "Feature count", "Value": feature_count},
    ]

    for class_name, count in (
        scenario_frame["ClassName"].value_counts(dropna=False).sort_index().items()
    ):
        rows.append(
            {
                "Metric": f"Class rows: {class_name}",
                "Value": int(count),
            }
        )

    for split_name in VALID_SPLITS:
        split_count = int((scenario_frame["Split"] == split_name).sum())
        rows.append(
            {
                "Metric": f"{split_name} rows",
                "Value": split_count,
            }
        )

        split_frame = scenario_frame[scenario_frame["Split"] == split_name]
        for class_name, count in (
            split_frame["ClassName"].value_counts(dropna=False).sort_index().items()
        ):
            rows.append(
                {
                    "Metric": f"{split_name} - {class_name}",
                    "Value": int(count),
                }
            )

    return pd.DataFrame(rows)


def audit_scenario(
    scenario_name: str,
    scenario_frame: pd.DataFrame,
    feature_columns: list[str],
    iqr_factor: float,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    إنشاء:
    - تقرير لكل خاصية.
    - تقرير القيم المتطرفة لكل خاصية ولكل Split.
    - أمثلة محدودة للقيم النصية وغير الصالحة.
    """
    train_frame = scenario_frame[scenario_frame["Split"] == "Train"].copy()
    if train_frame.empty:
        raise ValueError(f"{scenario_name}: Train split is empty.")

    feature_rows: list[dict[str, Any]] = []
    outlier_rows: list[dict[str, Any]] = []
    issue_examples: list[dict[str, Any]] = []

    max_issue_examples_per_feature = 10

    for feature in feature_columns:
        raw_all = scenario_frame[feature]
        raw_train = train_frame[feature]

        # القيم الفارغة الأصلية تشمل NaN والخلايا الفارغة النصية.
        blank_string_all = raw_all.astype("string").str.strip().eq("").fillna(False)
        blank_string_train = raw_train.astype("string").str.strip().eq("").fillna(False)

        missing_all = raw_all.isna() | blank_string_all
        missing_train = raw_train.isna() | blank_string_train

        # التحويل الرقمي؛ القيم النصية غير القابلة للتحويل تصبح NaN.
        numeric_all = pd.to_numeric(raw_all, errors="coerce")
        numeric_train = pd.to_numeric(raw_train, errors="coerce")

        text_mask_all = (~missing_all) & numeric_all.isna()
        text_mask_train = (~missing_train) & numeric_train.isna()

        pos_inf_all = np.isposinf(numeric_all.to_numpy(dtype=float, na_value=np.nan))
        neg_inf_all = np.isneginf(numeric_all.to_numpy(dtype=float, na_value=np.nan))
        finite_all = np.isfinite(numeric_all.to_numpy(dtype=float, na_value=np.nan))

        pos_inf_train = np.isposinf(
            numeric_train.to_numpy(dtype=float, na_value=np.nan)
        )
        neg_inf_train = np.isneginf(
            numeric_train.to_numpy(dtype=float, na_value=np.nan)
        )
        finite_train = np.isfinite(
            numeric_train.to_numpy(dtype=float, na_value=np.nan)
        )

        finite_values_all = numeric_all[finite_all]
        finite_values_train = numeric_train[finite_train]

        nonnegative_expected = expected_nonnegative(feature)
        negative_invalid_all = (
            int((finite_values_all < 0).sum()) if nonnegative_expected else 0
        )
        negative_invalid_train = (
            int((finite_values_train < 0).sum()) if nonnegative_expected else 0
        )

        cronbach_negative_all = (
            int((finite_values_all < 0).sum()) if is_cronbach_feature(feature) else 0
        )
        cronbach_negative_train = (
            int((finite_values_train < 0).sum()) if is_cronbach_feature(feature) else 0
        )

        unique_train = int(finite_values_train.nunique(dropna=True))
        constant_train = unique_train <= 1

        minimum_train = (
            float(finite_values_train.min()) if not finite_values_train.empty else np.nan
        )
        maximum_train = (
            float(finite_values_train.max()) if not finite_values_train.empty else np.nan
        )
        mean_train = (
            float(finite_values_train.mean()) if not finite_values_train.empty else np.nan
        )
        median_train = (
            float(finite_values_train.median()) if not finite_values_train.empty else np.nan
        )
        std_train = (
            float(finite_values_train.std(ddof=1))
            if len(finite_values_train) > 1
            else np.nan
        )

        if not finite_values_train.empty:
            q1_train = float(finite_values_train.quantile(0.25))
            q3_train = float(finite_values_train.quantile(0.75))
            iqr_train = q3_train - q1_train
            lower_bound = q1_train - iqr_factor * iqr_train
            upper_bound = q3_train + iqr_factor * iqr_train
        else:
            q1_train = q3_train = iqr_train = lower_bound = upper_bound = np.nan

        # في حال IQR = 0، لا نعتبر كل اختلاف Outlier تلقائيًا.
        valid_outlier_bounds = (
            np.isfinite(lower_bound)
            and np.isfinite(upper_bound)
            and iqr_train > 0
        )

        feature_row: dict[str, Any] = {
            "Scenario": scenario_name,
            "Feature": feature,
            "Expected_Nonnegative": nonnegative_expected,
            "Total_Rows": len(scenario_frame),
            "Train_Rows": len(train_frame),
            "Missing_All": int(missing_all.sum()),
            "Missing_Train": int(missing_train.sum()),
            "Text_All": int(text_mask_all.sum()),
            "Text_Train": int(text_mask_train.sum()),
            "Positive_Inf_All": int(pos_inf_all.sum()),
            "Positive_Inf_Train": int(pos_inf_train.sum()),
            "Negative_Inf_All": int(neg_inf_all.sum()),
            "Negative_Inf_Train": int(neg_inf_train.sum()),
            "Invalid_Negative_All": negative_invalid_all,
            "Invalid_Negative_Train": negative_invalid_train,
            "Cronbach_Negative_All_Allowed": cronbach_negative_all,
            "Cronbach_Negative_Train_Allowed": cronbach_negative_train,
            "Unique_Finite_Train": unique_train,
            "Constant_In_Train": constant_train,
            "Min_Train": minimum_train,
            "Max_Train": maximum_train,
            "Mean_Train": mean_train,
            "Median_Train": median_train,
            "Std_Train": std_train,
            "Q1_Train": q1_train,
            "Q3_Train": q3_train,
            "IQR_Train": iqr_train,
            "Outlier_Lower_Train": lower_bound,
            "Outlier_Upper_Train": upper_bound,
        }

        total_outliers = 0
        for split_name in VALID_SPLITS:
            split_mask = scenario_frame["Split"] == split_name
            split_values = numeric_all[split_mask]
            split_finite = split_values[np.isfinite(split_values)]

            if valid_outlier_bounds:
                lower_count = int((split_finite < lower_bound).sum())
                upper_count = int((split_finite > upper_bound).sum())
            else:
                lower_count = 0
                upper_count = 0

            outlier_count = lower_count + upper_count
            total_outliers += outlier_count

            outlier_rows.append(
                {
                    "Scenario": scenario_name,
                    "Feature": feature,
                    "Split": split_name,
                    "Rows_In_Split": int(split_mask.sum()),
                    "Finite_Values": int(len(split_finite)),
                    "Q1_From_Train": q1_train,
                    "Q3_From_Train": q3_train,
                    "IQR_From_Train": iqr_train,
                    "Lower_Bound_From_Train": lower_bound,
                    "Upper_Bound_From_Train": upper_bound,
                    "Lower_Outliers": lower_count,
                    "Upper_Outliers": upper_count,
                    "Total_Outliers": outlier_count,
                    "Outlier_Percent": (
                        100.0 * outlier_count / len(split_finite)
                        if len(split_finite)
                        else 0.0
                    ),
                    "Bounds_Usable": valid_outlier_bounds,
                }
            )

        feature_row["Outliers_All_Splits"] = total_outliers
        feature_rows.append(feature_row)

        # أمثلة محدودة فقط لتجنب إنشاء ملف ضخم.
        issue_mask = (
            missing_all
            | text_mask_all
            | pd.Series(pos_inf_all, index=scenario_frame.index)
            | pd.Series(neg_inf_all, index=scenario_frame.index)
        )
        if nonnegative_expected:
            issue_mask = issue_mask | (
                numeric_all.lt(0) & np.isfinite(numeric_all)
            )

        issue_indices = scenario_frame.index[issue_mask][
            :max_issue_examples_per_feature
        ]
        for row_index in issue_indices:
            raw_value = scenario_frame.at[row_index, feature]
            numeric_value = numeric_all.at[row_index]

            if pd.isna(raw_value) or str(raw_value).strip() == "":
                issue_type = "Missing"
            elif pd.isna(numeric_value):
                issue_type = "Text_or_NonNumeric"
            elif np.isposinf(numeric_value):
                issue_type = "Positive_Infinity"
            elif np.isneginf(numeric_value):
                issue_type = "Negative_Infinity"
            elif nonnegative_expected and numeric_value < 0:
                issue_type = "Invalid_Negative"
            else:
                issue_type = "Other"

            issue_examples.append(
                {
                    "Scenario": scenario_name,
                    "Feature": feature,
                    "Issue_Type": issue_type,
                    "ImgName": scenario_frame.at[row_index, "ImgName"]
                    if "ImgName" in scenario_frame.columns
                    else "",
                    "FileName": scenario_frame.at[row_index, "FileName"]
                    if "FileName" in scenario_frame.columns
                    else "",
                    "ClassName": scenario_frame.at[row_index, "ClassName"],
                    "Split": scenario_frame.at[row_index, "Split"],
                    "Raw_Value": safe_number(raw_value),
                }
            )

    feature_report = pd.DataFrame(feature_rows)
    outlier_report = pd.DataFrame(outlier_rows)
    issue_report = pd.DataFrame(issue_examples)

    # ترتيب الخصائص الأكثر حاجة للمراجعة أولًا.
    priority_columns = [
        "Missing_All",
        "Text_All",
        "Positive_Inf_All",
        "Negative_Inf_All",
        "Invalid_Negative_All",
        "Constant_In_Train",
        "Outliers_All_Splits",
    ]
    if not feature_report.empty:
        feature_report["Audit_Flag_Count"] = (
            feature_report["Missing_All"]
            + feature_report["Text_All"]
            + feature_report["Positive_Inf_All"]
            + feature_report["Negative_Inf_All"]
            + feature_report["Invalid_Negative_All"]
            + feature_report["Constant_In_Train"].astype(int)
        )
        feature_report = feature_report.sort_values(
            by=["Audit_Flag_Count", "Outliers_All_Splits", "Feature"],
            ascending=[False, False, True],
        )

    return feature_report, outlier_report, issue_report


def style_workbook(output_file: Path) -> None:
    """تنسيق بسيط وعملي للتقرير."""
    workbook = load_workbook(output_file)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F4CCCC")

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[1].height = 34

        # تحديد عرض الأعمدة بحد أقصى لمنع العرض المفرط.
        for column_cells in worksheet.columns:
            column_index = column_cells[0].column
            max_length = 0
            for cell in column_cells[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(max_length + 2, 10),
                32,
            )

        # تمييز الصفوف التي تحتوي مشكلات في أوراق Feature Audit.
        header_map = {
            cell.value: cell.column
            for cell in worksheet[1]
            if cell.value is not None
        }
        if "Audit_Flag_Count" in header_map:
            flag_column = header_map["Audit_Flag_Count"]
            for row_number in range(2, worksheet.max_row + 1):
                flag_value = worksheet.cell(row_number, flag_column).value
                if isinstance(flag_value, (int, float)) and flag_value > 0:
                    for cell in worksheet[row_number]:
                        cell.fill = error_fill

        if "Total_Outliers" in header_map:
            outlier_column = header_map["Total_Outliers"]
            for row_number in range(2, worksheet.max_row + 1):
                value = worksheet.cell(row_number, outlier_column).value
                if isinstance(value, (int, float)) and value > 0:
                    worksheet.cell(row_number, outlier_column).fill = warning_fill

    workbook.save(output_file)


def save_json_metadata(
    output_file: Path,
    input_file: Path,
    frame: pd.DataFrame,
    feature_columns: list[str],
    iqr_factor: float,
) -> None:
    metadata = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_file": str(input_file),
        "output_file": str(output_file),
        "rows": int(len(frame)),
        "columns": int(len(frame.columns)),
        "feature_columns": int(len(feature_columns)),
        "iqr_factor": float(iqr_factor),
        "scenarios": {key: list(value) for key, value in SCENARIOS.items()},
        "outlier_rule": "Q1 - factor*IQR, Q3 + factor*IQR; bounds fitted on Train only",
        "constant_rule": "Finite unique values in Train <= 1",
        "negative_rule": (
            "Negative values are flagged for all MC-CRoMD features except Cronbach."
        ),
        "python_version": platform.python_version(),
        "pandas_version": pd.__version__,
        "numpy_version": np.__version__,
    }

    metadata_path = output_file.with_name(f"{output_file.stem}_Metadata.json")
    metadata_path.write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Audit MC-CRoMD feature quality for three scenarios."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT_FILE,
        help="Features_With_Groups_And_Split.xlsx",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_FILE,
        help="Output Excel audit report.",
    )
    parser.add_argument(
        "--sheet",
        default=0,
        help="Input Excel sheet name or index.",
    )
    parser.add_argument(
        "--iqr-factor",
        type=float,
        default=DEFAULT_IQR_FACTOR,
    )

    # مناسب للتشغيل من Spyder.
    args, unknown_args = parser.parse_known_args()
    if unknown_args:
        print("Ignored unknown arguments:", unknown_args)

    input_file = Path(args.input)
    output_file = Path(args.output)

    print("Input file :", input_file)
    print("Output file:", output_file)
    print("IQR factor :", args.iqr_factor)

    if not input_file.exists():
        raise FileNotFoundError(f"Input file does not exist: {input_file}")

    if args.iqr_factor <= 0:
        raise ValueError("IQR factor must be greater than zero.")

    start_time = time.perf_counter()

    frame = read_input(input_file, sheet_name=args.sheet)
    frame.columns = [str(column).strip() for column in frame.columns]

    required_columns = {"ClassName", "Split"}
    missing_required = required_columns.difference(frame.columns)
    if missing_required:
        raise KeyError(
            f"Missing required columns: {sorted(missing_required)}"
        )

    frame["ClassName"] = frame["ClassName"].astype(str).str.strip()
    frame["Split"] = frame["Split"].map(normalize_split)

    unknown_splits = sorted(
        set(frame["Split"].dropna().unique()).difference(VALID_SPLITS)
    )
    if unknown_splits:
        raise ValueError(f"Unknown Split values: {unknown_splits}")

    feature_columns = detect_feature_columns(frame)
    print("Rows             :", len(frame))
    print("Feature columns  :", len(feature_columns))

    output_file.parent.mkdir(parents=True, exist_ok=True)

    overview = build_overview(
        frame=frame,
        feature_columns=feature_columns,
        input_file=input_file,
        iqr_factor=args.iqr_factor,
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="00_Overview", index=False)

        for scenario_name, scenario_classes in SCENARIOS.items():
            print(f"Auditing scenario: {scenario_name}")

            scenario_frame = frame[
                frame["ClassName"].isin(scenario_classes)
            ].copy()

            missing_classes = sorted(
                set(scenario_classes).difference(
                    set(scenario_frame["ClassName"].unique())
                )
            )
            if missing_classes:
                raise ValueError(
                    f"{scenario_name}: missing classes: {missing_classes}"
                )

            summary = make_scenario_summary(
                scenario_name=scenario_name,
                scenario_frame=scenario_frame,
                feature_count=len(feature_columns),
            )

            feature_report, outlier_report, issue_report = audit_scenario(
                scenario_name=scenario_name,
                scenario_frame=scenario_frame,
                feature_columns=feature_columns,
                iqr_factor=args.iqr_factor,
            )

            prefix = {
                "Binary": "Bin",
                "Three_Class": "Tri",
                "Five_Class": "Five",
            }[scenario_name]

            summary.to_excel(
                writer,
                sheet_name=f"{prefix}_Summary",
                index=False,
            )
            feature_report.to_excel(
                writer,
                sheet_name=f"{prefix}_Feature_Audit",
                index=False,
            )
            outlier_report.to_excel(
                writer,
                sheet_name=f"{prefix}_Outliers",
                index=False,
            )

            if issue_report.empty:
                issue_report = pd.DataFrame(
                    columns=[
                        "Scenario",
                        "Feature",
                        "Issue_Type",
                        "ImgName",
                        "FileName",
                        "ClassName",
                        "Split",
                        "Raw_Value",
                    ]
                )

            issue_report.to_excel(
                writer,
                sheet_name=f"{prefix}_Issue_Examples",
                index=False,
            )

    style_workbook(output_file)
    save_json_metadata(
        output_file=output_file,
        input_file=input_file,
        frame=frame,
        feature_columns=feature_columns,
        iqr_factor=args.iqr_factor,
    )

    elapsed = time.perf_counter() - start_time

    print("=" * 70)
    print("Audit completed successfully.")
    print("Excel report :", output_file)
    print(
        "Metadata     :",
        output_file.with_name(f"{output_file.stem}_Metadata.json"),
    )
    print(f"Execution time: {elapsed:.2f} seconds")


if __name__ == "__main__":
    main()
