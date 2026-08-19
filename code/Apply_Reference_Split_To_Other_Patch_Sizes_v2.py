# -*- coding: utf-8 -*-
"""
Apply the already-approved 256x256 group-aware split to another patch size.

Purpose
-------
- Use the completed 256-patch file as the ONLY reference for SourceGroupID/Split.
- Match the same images in the raw 128, 64, or 32 feature file by filename.
- Transfer the approved grouping/split columns without running GroupShuffleSplit again.
- Preserve target-file row order and all feature columns.
- Write an audit workbook and JSON metadata.

Designed for very wide Excel files and uses openpyxl streaming modes to reduce memory use.
Change only the three paths in the USER SETTINGS section.
"""

from __future__ import annotations

import hashlib
import json
import os
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


# =============================================================================
# USER SETTINGS: change only these paths for 128, 64, or 32
# =============================================================================

REFERENCE_256_FILE = Path(
    r"G:\My Research About Lung Canser\INASS\1-SourceGroupID"
    r"\GroupedSplitResults\256_5Clas\Features_With_Groups_And_Split.xlsx"
)

TARGET_RAW_FEATURE_FILE = Path(
    r"G:\My Research About Lung Canser\INASS\NewResults\32\32BZ_5Clas.xlsx"
)

OUTPUT_FILE = Path(
    r"G:\My Research About Lung Canser\INASS\1-SourceGroupID"
    r"\GroupedSplitResults\32_5Clas\Features_With_Groups_And_Split.xlsx"
)

# Stop if any image is unmatched or any class/label mismatch is found.
STRICT_MODE = True

# Preferred matching order.
# FileName is the primary key because LC25000 may contain different filenames
# with identical pixel content and therefore identical ImageSHA256 values.
# SHA256 remains available only as a fallback when it is unique.
MATCH_KEY_PRIORITY = ("FileName", "ImgName", "ImageSHA256")

# Columns copied from the approved 256 reference, when present.
TRANSFER_COLUMNS = (
    "SourceGroupID",
    "Split",
    "LC25000_GroupID",
    "LC25000_ClassName",
    "LC25000_FileName",
    "LC25000_Tissue",
    "LC25000_LocalCluster",
)

# Optional consistency checks between target and reference.
CHECK_COLUMNS = ("ClassName", "label")


# =============================================================================
# Helpers
# =============================================================================


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_filename(value: Any) -> str:
    text = normalize_text(value).replace("\\", "/")
    return text.rsplit("/", 1)[-1].lower()


def normalize_sha(value: Any) -> str:
    return normalize_text(value).lower()


def normalize_key(column: str, value: Any) -> str:
    if column == "ImageSHA256":
        return normalize_sha(value)
    if column in {"FileName", "ImgName"}:
        return normalize_filename(value)
    return normalize_text(value)


def file_sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while True:
            chunk = stream.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def get_header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(header_row):
        name = normalize_text(value)
        if not name:
            continue
        if name in mapping:
            raise ValueError(f"Duplicate column name found: {name}")
        mapping[name] = index
    return mapping


def choose_match_column(reference_headers: dict[str, int], target_headers: dict[str, int]) -> str:
    for candidate in MATCH_KEY_PRIORITY:
        if candidate in reference_headers and candidate in target_headers:
            return candidate
    raise KeyError(
        "No common matching column was found. Expected one of: "
        + ", ".join(MATCH_KEY_PRIORITY)
    )


def safe_equal(column: str, left: Any, right: Any) -> bool:
    if column == "label":
        try:
            return int(float(left)) == int(float(right))
        except Exception:
            return normalize_text(left) == normalize_text(right)
    return normalize_text(left).lower() == normalize_text(right).lower()


# =============================================================================
# Reference loading
# =============================================================================


def load_reference_lookup(reference_path: Path) -> tuple[
    dict[str, dict[str, Any]],
    dict[str, int],
    str,
    Counter,
]:
    """Read the approved 256 file once and build a lookup dictionary."""
    workbook = load_workbook(reference_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise ValueError("Reference file is empty.") from exc

    headers = get_header_map(header_row)

    missing_transfer = [
        column for column in ("SourceGroupID", "Split") if column not in headers
    ]
    if missing_transfer:
        raise KeyError(
            "Reference file is missing required columns: "
            + ", ".join(missing_transfer)
        )

    # Match column is selected later once target headers are known.
    # Here we collect all usable keys so fallback is possible.
    usable_key_columns = [column for column in MATCH_KEY_PRIORITY if column in headers]
    if not usable_key_columns:
        raise KeyError(
            "Reference file has none of the matching columns: "
            + ", ".join(MATCH_KEY_PRIORITY)
        )

    records_by_key_column: dict[str, dict[str, dict[str, Any]]] = {
        column: {} for column in usable_key_columns
    }
    split_counts: Counter = Counter()
    row_count = 0

    for row_number, row in enumerate(rows, start=2):
        row_count += 1
        record: dict[str, Any] = {}

        for column in TRANSFER_COLUMNS:
            if column in headers:
                record[column] = row[headers[column]]

        for column in CHECK_COLUMNS:
            if column in headers:
                record[column] = row[headers[column]]

        split_value = normalize_text(record.get("Split"))
        split_counts[split_value] += 1

        for key_column in usable_key_columns:
            key = normalize_key(key_column, row[headers[key_column]])
            if not key:
                continue
            lookup = records_by_key_column[key_column]
            if key in lookup:
                if key_column == "ImageSHA256":
                    # Different LC25000 filenames may contain identical image bytes.
                    # Such a SHA is ambiguous as a matching key, so remove it from
                    # the SHA lookup rather than treating it as a data error.
                    lookup[key] = None
                    continue

                raise ValueError(
                    f"Duplicate reference key in {key_column}: {key!r} "
                    f"(detected near row {row_number})"
                )

            lookup[key] = record

    workbook.close()

    # Remove ambiguous SHA keys created by byte-identical images with different names.
    if "ImageSHA256" in records_by_key_column:
        records_by_key_column["ImageSHA256"] = {
            key: value
            for key, value in records_by_key_column["ImageSHA256"].items()
            if value is not None
        }

    # Return all lookups inside one object; final column selected after target header read.
    return records_by_key_column, headers, str(row_count), split_counts


# =============================================================================
# Main transfer
# =============================================================================


def apply_reference_split() -> None:
    start_time = time.perf_counter()

    for path, label in (
        (REFERENCE_256_FILE, "Reference 256 file"),
        (TARGET_RAW_FEATURE_FILE, "Target raw feature file"),
    ):
        if not path.exists():
            raise FileNotFoundError(f"{label} does not exist:\n{path}")

    if REFERENCE_256_FILE.resolve() == TARGET_RAW_FEATURE_FILE.resolve():
        raise ValueError("Reference and target files must be different.")

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    audit_file = OUTPUT_FILE.with_name(f"{OUTPUT_FILE.stem}_ApplySplit_Audit.xlsx")
    metadata_file = OUTPUT_FILE.with_name(f"{OUTPUT_FILE.stem}_ApplySplit_Metadata.json")

    print("=" * 78)
    print("Apply approved 256 split to another patch size")
    print("=" * 78)
    print("Reference:", REFERENCE_256_FILE)
    print("Target   :", TARGET_RAW_FEATURE_FILE)
    print("Output   :", OUTPUT_FILE)
    print()

    reference_lookups, reference_headers, reference_row_count, reference_split_counts = (
        load_reference_lookup(REFERENCE_256_FILE)
    )

    target_workbook = load_workbook(
        TARGET_RAW_FEATURE_FILE,
        read_only=True,
        data_only=True,
    )
    target_sheet = target_workbook.active
    target_rows = target_sheet.iter_rows(values_only=True)

    try:
        target_header_row = next(target_rows)
    except StopIteration as exc:
        raise ValueError("Target feature file is empty.") from exc

    target_headers = get_header_map(target_header_row)
    match_column = choose_match_column(reference_headers, target_headers)
    reference_lookup = reference_lookups[match_column]

    print("Matching column:", match_column)
    print("Reference rows :", f"{len(reference_lookup):,}")

    # Preserve the provisional SourceGroupID under a clearer name.
    output_headers = [normalize_text(value) for value in target_header_row]

    if "SourceGroupID" in target_headers:
        source_index = target_headers["SourceGroupID"]
        if "Previous_SourceGroupID" not in target_headers:
            output_headers[source_index] = "Previous_SourceGroupID"
        else:
            # If both exist, remove the old SourceGroupID value later by replacing it.
            output_headers[source_index] = "Previous_SourceGroupID_From_Pre1"

    # Transfer only columns not already retained under their final names.
    transferable_present = [
        column for column in TRANSFER_COLUMNS if column in reference_headers
    ]

    for column in transferable_present:
        if column in output_headers:
            # Replace existing non-provisional values by writing into their current position.
            continue
        output_headers.append(column)

    output_header_map = {name: index for index, name in enumerate(output_headers)}

    output_workbook = Workbook(write_only=True)
    output_sheet = output_workbook.create_sheet("Features_With_Groups_And_Split")
    output_sheet.append(output_headers)

    unmatched_rows: list[dict[str, Any]] = []
    mismatch_rows: list[dict[str, Any]] = []
    duplicate_target_keys: list[dict[str, Any]] = []
    seen_target_keys: set[str] = set()
    transferred_split_counts: Counter = Counter()

    target_count = 0
    matched_count = 0

    for row_number, row in enumerate(target_rows, start=2):
        target_count += 1
        row_values = list(row)

        # Pad row to original header width if Excel omitted trailing empty cells.
        if len(row_values) < len(target_header_row):
            row_values.extend([None] * (len(target_header_row) - len(row_values)))

        key_raw = row[target_headers[match_column]]
        key = normalize_key(match_column, key_raw)

        if not key or key not in reference_lookup:
            unmatched_rows.append(
                {
                    "Target_Row": row_number,
                    "Match_Column": match_column,
                    "Match_Value": key_raw,
                    "FileName": row[target_headers["FileName"]]
                    if "FileName" in target_headers
                    else None,
                    "ImgName": row[target_headers["ImgName"]]
                    if "ImgName" in target_headers
                    else None,
                    "ClassName": row[target_headers["ClassName"]]
                    if "ClassName" in target_headers
                    else None,
                    "label": row[target_headers["label"]]
                    if "label" in target_headers
                    else None,
                }
            )
            record = None
        else:
            record = reference_lookup[key]
            matched_count += 1

            if key in seen_target_keys:
                duplicate_target_keys.append(
                    {
                        "Target_Row": row_number,
                        "Match_Column": match_column,
                        "Match_Value": key,
                    }
                )
            seen_target_keys.add(key)

            for check_column in CHECK_COLUMNS:
                if check_column in target_headers and check_column in record:
                    target_value = row[target_headers[check_column]]
                    reference_value = record[check_column]
                    if not safe_equal(check_column, target_value, reference_value):
                        mismatch_rows.append(
                            {
                                "Target_Row": row_number,
                                "Match_Value": key,
                                "Column": check_column,
                                "Target_Value": target_value,
                                "Reference_Value": reference_value,
                            }
                        )

        # Extend to final output width.
        if len(row_values) < len(output_headers):
            row_values.extend([None] * (len(output_headers) - len(row_values)))

        if record is not None:
            for column in transferable_present:
                output_index = output_header_map[column]
                row_values[output_index] = record.get(column)

            transferred_split_counts[normalize_text(record.get("Split"))] += 1

        output_sheet.append(row_values)

        if target_count % 250 == 0:
            print(
                f"Processed {target_count:,} rows | "
                f"Matched {matched_count:,} | Unmatched {len(unmatched_rows):,}"
            )

    target_workbook.close()

    # Strict validation before saving the final output.
    errors: list[str] = []
    if unmatched_rows:
        errors.append(f"Unmatched target rows: {len(unmatched_rows):,}")
    if mismatch_rows:
        errors.append(f"Class/label mismatches: {len(mismatch_rows):,}")
    if duplicate_target_keys:
        errors.append(f"Duplicate target match keys: {len(duplicate_target_keys):,}")
    if matched_count != target_count:
        errors.append(
            f"Matched count ({matched_count:,}) does not equal target count ({target_count:,})"
        )

    # Audit workbook is always saved, even when strict validation fails.
    audit_workbook = Workbook()
    summary_sheet = audit_workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Item", "Value"])
    summary_items = [
        ("Generated at", datetime.now().isoformat(timespec="seconds")),
        ("Reference file", str(REFERENCE_256_FILE)),
        ("Target file", str(TARGET_RAW_FEATURE_FILE)),
        ("Output file", str(OUTPUT_FILE)),
        ("Match column", match_column),
        ("Reference rows", int(reference_row_count)),
        ("Target rows", target_count),
        ("Matched rows", matched_count),
        ("Unmatched rows", len(unmatched_rows)),
        ("Class/label mismatches", len(mismatch_rows)),
        ("Duplicate target keys", len(duplicate_target_keys)),
        ("Strict mode", STRICT_MODE),
    ]
    for split_name in ("Train", "Validation", "Test"):
        summary_items.append(
            (f"Transferred {split_name} rows", transferred_split_counts[split_name])
        )
    for item, value in summary_items:
        summary_sheet.append([item, value])

    unmatched_sheet = audit_workbook.create_sheet("Unmatched")
    unmatched_headers = [
        "Target_Row",
        "Match_Column",
        "Match_Value",
        "FileName",
        "ImgName",
        "ClassName",
        "label",
    ]
    unmatched_sheet.append(unmatched_headers)
    for item in unmatched_rows:
        unmatched_sheet.append([item.get(column) for column in unmatched_headers])

    mismatch_sheet = audit_workbook.create_sheet("Class_Label_Mismatch")
    mismatch_headers = [
        "Target_Row",
        "Match_Value",
        "Column",
        "Target_Value",
        "Reference_Value",
    ]
    mismatch_sheet.append(mismatch_headers)
    for item in mismatch_rows:
        mismatch_sheet.append([item.get(column) for column in mismatch_headers])

    duplicate_sheet = audit_workbook.create_sheet("Duplicate_Target_Keys")
    duplicate_headers = ["Target_Row", "Match_Column", "Match_Value"]
    duplicate_sheet.append(duplicate_headers)
    for item in duplicate_target_keys:
        duplicate_sheet.append([item.get(column) for column in duplicate_headers])

    audit_workbook.save(audit_file)

    metadata = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "reference_file": str(REFERENCE_256_FILE),
        "target_file": str(TARGET_RAW_FEATURE_FILE),
        "output_file": str(OUTPUT_FILE),
        "audit_file": str(audit_file),
        "match_column": match_column,
        "reference_rows": int(reference_row_count),
        "target_rows": target_count,
        "matched_rows": matched_count,
        "unmatched_rows": len(unmatched_rows),
        "class_label_mismatches": len(mismatch_rows),
        "duplicate_target_keys": len(duplicate_target_keys),
        "reference_split_counts": dict(reference_split_counts),
        "transferred_split_counts": dict(transferred_split_counts),
        "transferred_columns": transferable_present,
        "policy": (
            "The 256x256 group-aware split is reused exactly. "
            "No random split is generated for the target patch size."
        ),
        "strict_mode": STRICT_MODE,
    }
    metadata_file.write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    if STRICT_MODE and errors:
        # Do not save a possibly invalid final feature file.
        print("\nValidation failed. Final output was not saved.")
        print("Audit file:", audit_file)
        for error in errors:
            print("-", error)
        raise RuntimeError(" | ".join(errors))

    output_workbook.save(OUTPUT_FILE)

    elapsed = time.perf_counter() - start_time
    print()
    print("=" * 78)
    print("Completed successfully")
    print("=" * 78)
    print("Target rows      :", f"{target_count:,}")
    print("Matched rows     :", f"{matched_count:,}")
    print("Unmatched rows   :", f"{len(unmatched_rows):,}")
    print("Mismatches       :", f"{len(mismatch_rows):,}")
    print("Duplicate keys   :", f"{len(duplicate_target_keys):,}")
    print("Split counts     :", dict(transferred_split_counts))
    print("Output file      :", OUTPUT_FILE)
    print("Audit file       :", audit_file)
    print("Metadata file    :", metadata_file)
    print(f"Execution seconds: {elapsed:.2f}")


if __name__ == "__main__":
    apply_reference_split()
