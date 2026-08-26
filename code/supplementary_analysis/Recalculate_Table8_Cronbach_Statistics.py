# -*- coding: utf-8 -*-
"""
Recalculate Table 8 - Cronbach's Alpha Statistics
MC-CRoMD
"""
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl

FILES = {
    "32x32":  r"G:\My Research About Lung Canser\INASS\1-SourceGroupID\GroupedSplitResults\32_5Clas\Features_With_Groups_And_Split.xlsx",
    "64x64":  r"G:\My Research About Lung Canser\INASS\1-SourceGroupID\GroupedSplitResults\64_5Clas\Features_With_Groups_And_Split.xlsx",
    "128x128": r"G:\My Research About Lung Canser\INASS\1-SourceGroupID\GroupedSplitResults\128_5Clas\Features_With_Groups_And_Split.xlsx",
    "256x256": r"G:\My Research About Lung Canser\INASS\1-SourceGroupID\GroupedSplitResults\256_5Clas\Features_With_Groups_And_Split.xlsx",
}

OUTPUT_FILE = r"G:\My Research About Lung Canser\INASS\Table8_Cronbach_Statistics_Recalculated.xlsx"

EXPECTED_CRONBACH_COLUMNS = {
    "32x32": 576,
    "64x64": 144,
    "128x128": 36,
    "256x256": 9,
}

def calculate_cronbach_statistics(file_path, patch_size):
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"File not found for {patch_size}:\n{file_path}")

    print("=" * 80)
    print(f"Patch size : {patch_size}")
    print(f"File       : {file_path}")
    print("=" * 80)

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))

    cronbach_indices, cronbach_names = [], []
    for index, column_name in enumerate(header_row):
        if isinstance(column_name, str) and column_name.startswith("Cronbach_"):
            cronbach_indices.append(index)
            cronbach_names.append(column_name)

    expected_count = EXPECTED_CRONBACH_COLUMNS[patch_size]
    print(f"Cronbach columns found : {len(cronbach_names)}")
    print(f"Expected columns       : {expected_count}")

    if len(cronbach_names) != expected_count:
        workbook.close()
        raise ValueError(
            f"Unexpected number of Cronbach columns for {patch_size}.\n"
            f"Expected: {expected_count}\nFound   : {len(cronbach_names)}"
        )

    valid_values = []
    invalid_count = 0

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        for column_index in cronbach_indices:
            value = row[column_index]
            if value is None:
                invalid_count += 1
                continue
            try:
                value = float(value)
            except (TypeError, ValueError):
                invalid_count += 1
                continue
            if not np.isfinite(value):
                invalid_count += 1
                continue
            valid_values.append(value)
        if row_number % 1000 == 0:
            print(f"Processed rows: {row_number - 1:,}", end="\r")

    workbook.close()
    values = np.asarray(valid_values, dtype=np.float64)
    if values.size == 0:
        raise ValueError(f"No valid Cronbach values found for {patch_size}.")

    mean_alpha = np.mean(values)
    std_alpha = np.std(values, ddof=1)
    median_alpha = np.median(values)

    print()
    print(f"Valid Alpha values : {values.size:,}")
    print(f"Invalid values     : {invalid_count:,}")
    print(f"Mean Alpha         : {mean_alpha:.10f}")
    print(f"Std Alpha          : {std_alpha:.10f}")
    print(f"Median Alpha       : {median_alpha:.10f}\n")

    return {
        "Patch Size": patch_size,
        "Cronbach Columns": len(cronbach_names),
        "Valid Alpha Values": int(values.size),
        "Invalid Values": int(invalid_count),
        "Mean Alpha": float(mean_alpha),
        "Std Alpha": float(std_alpha),
        "Median Alpha": float(median_alpha),
        "Mean Alpha (3 d.p.)": round(float(mean_alpha), 3),
        "Std Alpha (3 d.p.)": round(float(std_alpha), 3),
        "Median Alpha (3 d.p.)": round(float(median_alpha), 3),
    }

def main():
    print("\nMC-CRoMD - Recalculation of Table 8")
    print("Cronbach's Alpha descriptive statistics\n")

    results = []
    for patch_size, file_path in FILES.items():
        if "PATH_TO_" in file_path:
            print(f"{patch_size}: path has not been specified yet — skipped.")
            continue
        results.append(calculate_cronbach_statistics(file_path, patch_size))

    if not results:
        print("\nNo files were processed.")
        print("Please insert the correct file paths in FILES.")
        return

    results_df = pd.DataFrame(results)
    patch_order = ["32x32", "64x64", "128x128", "256x256"]
    results_df["Patch_Order"] = results_df["Patch Size"].map({p: i for i, p in enumerate(patch_order)})
    results_df = results_df.sort_values("Patch_Order").drop(columns=["Patch_Order"]).reset_index(drop=True)

    manuscript_table = results_df[[
        "Patch Size", "Mean Alpha (3 d.p.)", "Std Alpha (3 d.p.)", "Median Alpha (3 d.p.)"
    ]].copy()
    manuscript_table.columns = ["Patch Size", "Mean α", "Std", "Median α"]

    output_path = Path(OUTPUT_FILE)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        manuscript_table.to_excel(writer, sheet_name="Table8_Manuscript", index=False)
        results_df.to_excel(writer, sheet_name="Detailed_Statistics", index=False)

    print("=" * 80)
    print("FINAL TABLE 8")
    print("=" * 80)
    print(manuscript_table.to_string(index=False))
    print("\nSaved to:")
    print(output_path)
    print("\nTABLE 8 RECALCULATION COMPLETED SUCCESSFULLY")

if __name__ == "__main__":
    main()
