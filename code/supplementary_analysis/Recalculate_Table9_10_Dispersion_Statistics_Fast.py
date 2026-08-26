# -*- coding: utf-8 -*-
"""
Recalculate Tables 9 and 10 - Dispersion Statistics
MC-CRoMD

Purpose
-------
Recalculate the manuscript values for:
    Table 9 : Range, Variance, STD, CV
    Table 10: Q1, Q3, IQR, MAD

Design goals
------------
1) Exact calculation from the stored MC-CRoMD feature files.
2) One streaming pass per XLSX file.
3) Very low RAM usage: values are NOT accumulated in large Python lists.
4) Strong header/count checks before calculation.
5) High-precision output + publication-ready rounded tables.
6) Internal consistency checks for Q1/Q3/IQR and Variance/STD.

Important
---------
This program verifies the values STORED in the feature files.
It can verify whether stored STD^2 agrees with stored Variance.
The choice of sample vs population variance (ddof=1 vs ddof=0) must
ultimately be confirmed from the original feature-extraction code.
"""

from pathlib import Path
import math
import re
import time
import numpy as np
import pandas as pd
import openpyxl


# ---------------------------------------------------------------------
# INPUT FILES
# ---------------------------------------------------------------------
FILES = {
    "32x32":  r"G:\My Research About Lung Canser\INASS\1-SourceGroupID\GroupedSplitResults\32_5Clas\Features_With_Groups_And_Split.xlsx",
    "64x64":  r"G:\My Research About Lung Canser\INASS\1-SourceGroupID\GroupedSplitResults\64_5Clas\Features_With_Groups_And_Split.xlsx",
    "128x128": r"G:\My Research About Lung Canser\INASS\1-SourceGroupID\GroupedSplitResults\128_5Clas\Features_With_Groups_And_Split.xlsx",
    "256x256": r"G:\My Research About Lung Canser\INASS\1-SourceGroupID\GroupedSplitResults\256_5Clas\Features_With_Groups_And_Split.xlsx",
}

OUTPUT_FILE = r"G:\My Research About Lung Canser\INASS\Table9_10_Dispersion_Statistics_Recalculated.xlsx"

PATCH_COUNTS = {
    "32x32": 576,
    "64x64": 144,
    "128x128": 36,
    "256x256": 9,
}

MEASURES = ("Range", "Variance", "STD", "Q1", "Q3", "IQR", "MAD", "CV")


# ---------------------------------------------------------------------
# HEADER CLASSIFICATION
# ---------------------------------------------------------------------
def normalize_name(name):
    """Normalize a column name for robust matching."""
    if not isinstance(name, str):
        return ""
    return re.sub(r"[^A-Z0-9]+", "", name.upper())


def classify_measure(column_name):
    """
    Return one of MEASURES if the column is a dispersion-feature column,
    otherwise return None.

    Matching is intentionally conservative. Cronbach columns are excluded.
    """
    s = normalize_name(column_name)
    if not s or "CRONBACH" in s:
        return None

    # Long/specific names first.
    if "INTERQUARTILERANGE" in s or "IQR" in s:
        return "IQR"
    if "MEANABSOLUTEDEVIATION" in s or s.endswith("MAD") or "MAD" in s:
        return "MAD"
    if "COEFFICIENTOFVARIATION" in s or s.endswith("CV") or "CV" in s:
        return "CV"
    if "STANDARDDEVIATION" in s or "STDDEV" in s or "STDEV" in s or "STD" in s:
        return "STD"
    if "VARIANCE" in s or s.endswith("VAR"):
        return "Variance"
    if "RANGE" in s:
        return "Range"

    # Quartiles.
    if "Q1" in s or "25THPERCENTILE" in s or "P25" in s:
        return "Q1"
    if "Q3" in s or "75THPERCENTILE" in s or "P75" in s:
        return "Q3"

    return None


def get_measure_columns(header_row, patch_size):
    """Identify dispersion columns and verify the expected count."""
    groups = {m: [] for m in MEASURES}

    for idx, name in enumerate(header_row):
        measure = classify_measure(name)
        if measure is not None:
            groups[measure].append(idx)

    expected_per_measure = PATCH_COUNTS[patch_size] * 3  # 3 RGB channels

    print("\nColumn verification")
    print("-" * 80)
    print(f"Expected columns per measure: {expected_per_measure:,}")

    bad = []
    for measure in MEASURES:
        found = len(groups[measure])
        print(f"{measure:10s}: {found:,}")
        if found != expected_per_measure:
            bad.append((measure, found))

    if bad:
        details = ", ".join(f"{m}={n:,}" for m, n in bad)
        raise ValueError(
            f"\nUnexpected dispersion-column count for {patch_size}: {details}\n"
            f"Expected {expected_per_measure:,} columns for EACH measure.\n"
            "The program stopped BEFORE calculating statistics so that no "
            "incorrect table can be produced."
        )

    total = sum(len(v) for v in groups.values())
    expected_total = expected_per_measure * len(MEASURES)

    if total != expected_total:
        raise ValueError(
            f"Unexpected total dispersion columns for {patch_size}. "
            f"Expected {expected_total:,}, found {total:,}."
        )

    print(f"Total dispersion columns verified: {total:,}")
    return groups


# ---------------------------------------------------------------------
# FAST STREAMING AGGREGATION
# ---------------------------------------------------------------------
def update_sum_count(values, total_sum, total_count, invalid_count):
    """
    Aggregate a sequence without storing it.
    Returns updated (sum, count, invalid).
    """
    local_sum = 0.0
    local_count = 0
    local_invalid = 0

    for value in values:
        if value is None:
            local_invalid += 1
            continue

        # Excel numeric cells normally arrive as int/float.
        if isinstance(value, (int, float, np.integer, np.floating)):
            x = float(value)
        else:
            try:
                x = float(value)
            except (TypeError, ValueError):
                local_invalid += 1
                continue

        if not math.isfinite(x):
            local_invalid += 1
            continue

        local_sum += x
        local_count += 1

    return (
        total_sum + local_sum,
        total_count + local_count,
        invalid_count + local_invalid,
    )


def calculate_file(file_path, patch_size):
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"File not found for {patch_size}:\n{file_path}")

    print("\n" + "=" * 80)
    print(f"Patch size : {patch_size}")
    print(f"File       : {file_path}")
    print("=" * 80)

    t0 = time.perf_counter()

    wb = openpyxl.load_workbook(
        file_path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    ws = wb[wb.sheetnames[0]]

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    groups = get_measure_columns(header, patch_size)

    # For speed, create one sorted list of only the required feature columns.
    # Each entry: (column_index, measure)
    selected = []
    for measure, indices in groups.items():
        selected.extend((idx, measure) for idx in indices)
    selected.sort(key=lambda x: x[0])

    sums = {m: 0.0 for m in MEASURES}
    counts = {m: 0 for m in MEASURES}
    invalid = {m: 0 for m in MEASURES}

    # Additional paired checks:
    # Q3 - Q1 versus IQR, when corresponding columns can be paired by order.
    # Variance versus STD^2, when corresponding columns can be paired by order.
    q1_idx = groups["Q1"]
    q3_idx = groups["Q3"]
    iqr_idx = groups["IQR"]
    var_idx = groups["Variance"]
    std_idx = groups["STD"]

    iqr_diff_sum = 0.0
    iqr_diff_count = 0
    var_std_diff_sum = 0.0
    var_std_diff_count = 0

    print("\nStreaming rows (single pass, low-memory)...")

    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        # Main table aggregates.
        # Using precomputed indices avoids repeated header/name work.
        for idx, measure in selected:
            value = row[idx]

            if value is None:
                invalid[measure] += 1
                continue

            if isinstance(value, (int, float, np.integer, np.floating)):
                x = float(value)
            else:
                try:
                    x = float(value)
                except (TypeError, ValueError):
                    invalid[measure] += 1
                    continue

            if not math.isfinite(x):
                invalid[measure] += 1
                continue

            sums[measure] += x
            counts[measure] += 1

        # Consistency check: mean absolute difference
        # IQR vs (Q3-Q1), paired by feature order.
        for a, b, c in zip(q1_idx, q3_idx, iqr_idx):
            q1 = row[a]
            q3 = row[b]
            iq = row[c]
            if all(isinstance(v, (int, float, np.integer, np.floating)) for v in (q1, q3, iq)):
                q1 = float(q1)
                q3 = float(q3)
                iq = float(iq)
                if math.isfinite(q1) and math.isfinite(q3) and math.isfinite(iq):
                    iqr_diff_sum += abs((q3 - q1) - iq)
                    iqr_diff_count += 1

        # Consistency check: Variance vs STD^2, paired by feature order.
        for a, b in zip(var_idx, std_idx):
            va = row[a]
            sd = row[b]
            if isinstance(va, (int, float, np.integer, np.floating)) and \
               isinstance(sd, (int, float, np.integer, np.floating)):
                va = float(va)
                sd = float(sd)
                if math.isfinite(va) and math.isfinite(sd):
                    var_std_diff_sum += abs(va - sd * sd)
                    var_std_diff_count += 1

        if row_no % 1000 == 0:
            elapsed = time.perf_counter() - t0
            print(
                f"Processed rows: {row_no - 1:,} | "
                f"Elapsed: {elapsed/60:.1f} min",
                end="\r",
            )

    wb.close()
    print()

    means = {}
    for measure in MEASURES:
        if counts[measure] == 0:
            raise ValueError(f"No valid values found for {measure} in {patch_size}.")
        means[measure] = sums[measure] / counts[measure]

    elapsed = time.perf_counter() - t0

    print("\nCalculated means")
    print("-" * 80)
    for measure in MEASURES:
        print(
            f"{measure:10s}: {means[measure]:.10f} "
            f"| valid={counts[measure]:,} | invalid={invalid[measure]:,}"
        )

    mean_iqr_identity_error = (
        iqr_diff_sum / iqr_diff_count if iqr_diff_count else float("nan")
    )
    mean_var_std2_error = (
        var_std_diff_sum / var_std_diff_count if var_std_diff_count else float("nan")
    )

    print("\nInternal checks")
    print("-" * 80)
    print(
        "Mean absolute error of stored IQR vs (Q3-Q1): "
        f"{mean_iqr_identity_error:.12g}"
    )
    print(
        "Mean absolute error of stored Variance vs STD^2: "
        f"{mean_var_std2_error:.12g}"
    )
    print(f"Elapsed for {patch_size}: {elapsed/60:.2f} minutes")

    result = {
        "Patch Size": patch_size,
        **{m: means[m] for m in MEASURES},
        **{f"{m} Valid": counts[m] for m in MEASURES},
        **{f"{m} Invalid": invalid[m] for m in MEASURES},
        "Mean |IQR-(Q3-Q1)|": mean_iqr_identity_error,
        "Mean |Variance-STD^2|": mean_var_std2_error,
        "Elapsed Minutes": elapsed / 60.0,
    }
    return result


# ---------------------------------------------------------------------
# OUTPUT
# ---------------------------------------------------------------------
def main():
    print("\nMC-CRoMD - Recalculation of Tables 9 and 10")
    print("Dispersion statistics verification")
    print("Exact stored-feature aggregation; one streaming pass per file.\n")

    all_results = []

    for patch_size in ("32x32", "64x64", "128x128", "256x256"):
        all_results.append(calculate_file(FILES[patch_size], patch_size))

    df = pd.DataFrame(all_results)

    # Publication-ready tables: 4 decimal places for consistency.
    table9 = df[["Patch Size", "Range", "Variance", "STD", "CV"]].copy()
    table10 = df[["Patch Size", "Q1", "Q3", "IQR", "MAD"]].copy()

    for col in table9.columns[1:]:
        table9[col] = table9[col].round(4)
    for col in table10.columns[1:]:
        table10[col] = table10[col].round(4)

    print("\n" + "=" * 80)
    print("RECALCULATED TABLE 9")
    print("=" * 80)
    print(table9.to_string(index=False))

    print("\n" + "=" * 80)
    print("RECALCULATED TABLE 10")
    print("=" * 80)
    print(table10.to_string(index=False))

    output_path = Path(OUTPUT_FILE)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        table9.to_excel(writer, sheet_name="Table9_Manuscript", index=False)
        table10.to_excel(writer, sheet_name="Table10_Manuscript", index=False)

        df[[
            "Patch Size", "Range", "Variance", "STD", "CV",
            "Q1", "Q3", "IQR", "MAD",
            "Mean |IQR-(Q3-Q1)|",
            "Mean |Variance-STD^2|",
            "Elapsed Minutes",
        ]].to_excel(writer, sheet_name="Detailed_Statistics", index=False)

        validity_cols = ["Patch Size"]
        for m in MEASURES:
            validity_cols.extend([f"{m} Valid", f"{m} Invalid"])
        df[validity_cols].to_excel(
            writer, sheet_name="Validity_Counts", index=False
        )

    print("\nSaved to:")
    print(output_path)
    print("\nTABLES 9 AND 10 RECALCULATION COMPLETED SUCCESSFULLY")
    print("\nImportant:")
    print(
        "This verifies the stored feature values. "
        "The sample/population variance denominator must be confirmed "
        "from the original feature-extraction code."
    )


if __name__ == "__main__":
    main()
